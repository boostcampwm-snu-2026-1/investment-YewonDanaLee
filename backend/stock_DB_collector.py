import os
from datetime import datetime, timedelta
import openpyxl
from pykrx import stock

# 종목 코드 매핑 (이름 → KRX 티커)
STOCKS = {
    "삼성": "005930",   # 삼성전자
    "SK":   "000660",   # SK하이닉스
}

def get_latest_trading_day() -> str:
    """오늘이 장 마감 후면 오늘, 아니면 어제 (주말/공휴일 포함 역추적)"""
    now = datetime.now()
    # 한국 장 마감 = 15:30. 그 전이면 전날 데이터가 최신
    if now.hour < 15 or (now.hour == 15 and now.minute < 30):
        now -= timedelta(days=1)
    # 주말 스킵
    while now.weekday() >= 5:
        now -= timedelta(days=1)
    return now.strftime("%Y%m%d")

def fetch_stock_data(ticker: str, base_date: str) -> dict | None:
    """
    pykrx로 OHLCV + 변동% 계산
    - base_date  : 저장할 날짜 (YYYYMMDD)
    - 변동%      : (당일종가 - 전일종가) / 전일종가 * 100
    """
    # 전일 포함 2거래일치 가져오기 (변동% 계산용)
    start = (datetime.strptime(base_date, "%Y%m%d") - timedelta(days=10)).strftime("%Y%m%d")

    df = stock.get_market_ohlcv_by_date(start, base_date, ticker)

    if df is None or df.empty:
        print(f"    [실패] 데이터 없음 (ticker={ticker}, date={base_date})")
        return None

    # 마지막 행 = base_date 당일 데이터
    today_row = df.iloc[-1]

    # 변동% 계산
    if len(df) >= 2:
        prev_close = df.iloc[-2]["종가"]
        change_pct = (today_row["종가"] - prev_close) / prev_close * 100
        change_str = f"{change_pct:+.2f}%"
    else:
        change_str = "N/A"

    # 날짜를 "YYYY년 MM월 DD일" 형식으로 변환
    date_label = df.index[-1].strftime("%Y년 %m월 %d일")

    # 거래량 포맷 (17,180,000 → "17.18M")
    volume = today_row["거래량"]
    if volume >= 1_000_000:
        volume_str = f"{volume / 1_000_000:.2f}M"
    elif volume >= 1_000:
        volume_str = f"{volume / 1_000:.1f}K"
    else:
        volume_str = str(volume)

    return {
        "date":   date_label,
        "close":  f"{int(today_row['종가']):,}",
        "open":   f"{int(today_row['시가']):,}",
        "high":   f"{int(today_row['고가']):,}",
        "low":    f"{int(today_row['저가']):,}",
        "volume": volume_str,
        "change": change_str,
    }

def save_to_excel(prefix: str, data: dict):
    data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data")
    os.makedirs(data_dir, exist_ok=True)
    filename = os.path.join(data_dir, f"{prefix}_history.xlsx")

    # 파일 없으면 헤더 포함 새로 생성
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "시트1"
        ws.append(["날짜", "종가", "시가", "고가", "저가", "거래량", "변동 %"])
        ws.column_dimensions["A"].width = 18   # 날짜 열 ### 방지
        wb.save(filename)

    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # 날짜 열 너비 보정 (기존 파일 대응)
    ws.column_dimensions["A"].width = 18

    # 중복 저장 방지
    if ws.max_row >= 2:
        existing_date = ws.cell(row=2, column=1).value
        if existing_date == data["date"]:
            print(f"  👉 [스킵] {data['date']} 데이터가 이미 존재합니다.\n")
            wb.close()
            return

    ws.insert_rows(2)
    ws.cell(row=2, column=1, value=data["date"])
    ws.cell(row=2, column=2, value=data["close"])
    ws.cell(row=2, column=3, value=data["open"])
    ws.cell(row=2, column=4, value=data["high"])
    ws.cell(row=2, column=5, value=data["low"])
    ws.cell(row=2, column=6, value=data["volume"])
    ws.cell(row=2, column=7, value=data["change"])

    wb.save(filename)
    wb.close()
    print(f"  👉 [성공] {data['date']} → {filename} 2행에 저장 완료\n")

def main():
    print("[시작] 국내 주식 전일 마감 데이터 수집 (pykrx)\n")

    base_date = get_latest_trading_day()
    print(f"  기준 날짜: {base_date}\n")

    for prefix, ticker in STOCKS.items():
        print(f"[{prefix}_history] 수집 중... (ticker={ticker})")
        data = fetch_stock_data(ticker, base_date)
        if data:
            print(f"    {data['date']} | 종가 {data['close']} | 변동 {data['change']}")
            save_to_excel(prefix, data)
        else:
            print(f"    [실패] {prefix} 데이터를 가져오지 못했습니다.\n")

    print("[완료] 모든 수집 프로세스 종료.")

if __name__ == "__main__":
    main()