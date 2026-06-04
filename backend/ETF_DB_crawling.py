import os
import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

# 인베스팅닷컴 타겟 URL 딕셔너리
URLS = {
    "SOX": "https://kr.investing.com/indices/phlx-semiconductor-historical-data",
    "DRAM_ETF": "https://kr.investing.com/etfs/dram-historical-data",
    "EWY": "https://kr.investing.com/etfs/ishares-south-korea-index-historical-data",
    "KORU": "https://kr.investing.com/etfs/direxion-daily-sk-bull-3x-shrs-historical-data"
}

EXCEL_NAME = "ETF.xlsx"

def create_driver() -> webdriver.Chrome:
    """우회용 크롬 드라이버 설정"""
    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    )
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    
    driver = webdriver.Chrome(options=opts)
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"},
    )
    return driver


def fetch_latest_row(driver, url):
    """지정한 과거 데이터 페이지에서 가장 최신 날짜(첫 번째 행)와 종가를 파싱"""
    driver.get(url)
    try:
        # 과거 데이터 테이블 로딩 대기
        table = WebDriverWait(driver, 12).until(
            EC.presence_of_element_located((By.XPATH, '//table[contains(@class, "freeze-column")] | //table'))
        )
        
        # 첫 번째 데이터 행(tbody -> 첫 번째 tr) 조준
        first_tr = table.find_element(By.XPATH, './/tbody/tr[1]')
        
        # 날짜 추출 (<time> 태그 타겟팅)
        time_el = first_tr.find_element(By.TAG_NAME, 'time')
        date_str = time_el.text.strip() # 예: "2026년 06월 03일"
        
        # 종가(Close Price) 추출: 날짜 바로 다음 열(td)이 보통 종가입니다.
        td_elements = first_tr.find_elements(By.TAG_NAME, 'td')
        
        # 첫 번째 td가 날짜를 포함하거나 비어있을 수 있으므로 안전하게 종가 텍스트 확보
        # 인베스팅닷컴 신형 테이블 기준 2번째(인덱스 1)가 종가입니다.
        close_price_raw = td_elements[1].text.strip()
        
        # 쉼표를 제거하고 실수(float)로 변환
        close_price = float(close_price_raw.replace(",", ""))
        
        return date_str, close_price
    except Exception as e:
        print(f"[에러] {url} 파싱 실패: {e}")
        return None, None


def main():
    driver = create_driver()
    scraped_data = {}
    common_date = None
    
    print("[시작] 인베스팅닷컴 과거 데이터 수집을 시작합니다...")
    
    for key, url in URLS.items():
        print(f"  - {key} 데이터 가져오는 중...")
        date_str, price = fetch_latest_row(driver, url)
        
        if date_str and price:
            scraped_data[key] = price
            # 모든 자산의 최신 기준 날짜가 동일한지 체크 (일반적으로 미국 장 마감 기준 동일)
            if not common_date:
                common_date = date_str
        else:
            print(f"[경고] {key} 데이터를 수집하지 못해 작업을 중단합니다.")
            driver.quit()
            return

    driver.quit()
    
    if not common_date or len(scraped_data) < 4:
        print("[종료] 완전한 데이터를 수집하지 못했습니다.")
        return

    print(f"\n[수집 성공] 기준 날짜: {common_date}")
    for k, v in scraped_data.items():
        print(f"  * {k}: {v:,}")

    # ----------------------------------------------------
    # 3. 엑셀 파일 처리 및 중복 체크 검증
    # ----------------------------------------------------
    # 엑셀 파일이 없으면 새로 생성 (헤더 포함)
    if not os.path.exists(EXCEL_NAME):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "시트1"
        ws.append(["날짜", "SOX", "DRAM_ETF", "EWY", "KORU"])
        wb.save(EXCEL_NAME)

    # 엑셀 로드
    wb = openpyxl.load_workbook(EXCEL_NAME)
    ws = wb.active

    # 기존 데이터가 있을 경우 2행(기존의 가장 최신 행)의 날짜와 비교하여 중복 체크
    if ws.max_row >= 2:
        existing_latest_date = ws.cell(row=2, column=1).value
        if existing_latest_date == common_date:
            print(f"\n[스킵] 최신 날짜 '{common_date}' 데이터가 이미 엑셀에 존재합니다. 크롤링을 스킵합니다.")
            wb.close()
            return

    # 💡 핵심: 2행에 빈 행을 삽입하여 기존 데이터를 아래로 밀어냅니다.
    ws.insert_rows(2)
    
    # 새 데이터 삽입 (2행)
    ws.cell(row=2, column=1, value=common_date)
    ws.cell(row=2, column=2, value=scraped_data["SOX"])
    ws.cell(row=2, column=3, value=scraped_data["DRAM_ETF"])
    ws.cell(row=2, column=4, value=scraped_data["EWY"])
    ws.cell(row=2, column=5, value=scraped_data["KORU"])

    # 엑셀 저장 및 닫기
    wb.save(EXCEL_NAME)
    wb.close()
    print(f"\n[완료] '{common_date}' 행이 {EXCEL_NAME}의 2행에 성공적으로 추가되었습니다!")

if __name__ == "__main__":
    main()