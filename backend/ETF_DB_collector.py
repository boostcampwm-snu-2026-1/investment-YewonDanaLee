import os
from datetime import datetime, timedelta
import yfinance as yf
import openpyxl

# 이름 → Yahoo Finance 티커
TICKERS = {
    "SOX":      "^SOX",   # PHLX 반도체 지수
    "DRAM_ETF": "DRAM",   # DRAM ETF
    "EWY":      "EWY",    # iShares MSCI South Korea ETF
    "KORU":     "KORU",   # Direxion Daily South Korea Bull 3X
}

EXCEL_NAME = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "ETF.xlsx")
KEYS = ["SOX", "DRAM_ETF", "EWY", "KORU"]


def get_latest_trading_day() -> datetime:
    """미국 기준: 오늘이 주말이면 가장 최근 금요일로 역추적"""
    now = datetime.now()
    while now.weekday() >= 5:   # 5=토, 6=일
        now -= timedelta(days=1)
    return now


def fetch_all() -> tuple[str | None, dict]:
    """
    yfinance로 전 종목 일괄 다운로드.
    반환: (날짜 문자열, {키: 종가})
    """
    base_day = get_latest_trading_day()
    start = (base_day - timedelta(days=7)).strftime("%Y-%m-%d")   # 넉넉히 7일
    end   = (base_day + timedelta(days=1)).strftime("%Y-%m-%d")   # end는 exclusive

    symbols = list(TICKERS.values())
    # 종목이 여럿이면 group_by='ticker'로 멀티컬럼 DataFrame 반환
    raw = yf.download(symbols, start=start, end=end,
                      auto_adjust=True, progress=False, group_by="ticker")

    if raw.empty:
        print("[에러] yfinance 다운로드 결과가 비어있습니다.")
        return None, {}

    result = {}
    common_date = None

    for name, symbol in TICKERS.items():
        try:
            # 멀티티커일 때 컬럼 구조: (ticker, OHLCV)
            close_series = raw[symbol]["Close"].dropna()
            if close_series.empty:
                print(f"  [경고] {name}({symbol}): 데이터 없음")
                return None, {}

            last_date  = close_series.index[-1]
            last_close = float(close_series.iloc[-1])

            date_str = last_date.strftime("%Y년 %m월 %d일")
            if common_date is None:
                common_date = date_str
            elif common_date != date_str:
                # 종목마다 최신 거래일이 다를 수 있으면 경고만 출력
                print(f"  [주의] {name} 날짜({date_str})가 기준 날짜({common_date})와 다릅니다.")

            result[name] = last_close
            print(f"  - {name}({symbol}): {date_str} 종가={last_close:,.2f}")

        except KeyError:
            print(f"  [경고] {name}({symbol}): 컬럼 없음 — 상장폐지 또는 잘못된 티커")
            return None, {}

    return common_date, result


def save_to_excel(common_date: str, data: dict):
    os.makedirs(os.path.dirname(EXCEL_NAME), exist_ok=True)

    # 파일 없으면 헤더 포함 새로 생성
    if not os.path.exists(EXCEL_NAME):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "시트1"
        ws.append(["날짜"] + KEYS)
        ws.column_dimensions["A"].width = 18
        wb.save(EXCEL_NAME)

    wb = openpyxl.load_workbook(EXCEL_NAME)
    ws = wb.active
    ws.column_dimensions["A"].width = 18   # 기존 파일도 보정

    # 같은 날짜 행이 있으면 삭제 후 덮어쓰기
    if ws.max_row >= 2:
        to_delete = [r for r in range(2, ws.max_row + 1)
                     if ws.cell(row=r, column=1).value == common_date]
        for r in reversed(to_delete):
            ws.delete_rows(r)

    ws.insert_rows(2)
    ws.cell(row=2, column=1, value=common_date)
    for col_idx, key in enumerate(KEYS, start=2):
        ws.cell(row=2, column=col_idx, value=data[key])

    wb.save(EXCEL_NAME)
    wb.close()
    print(f"\n[완료] '{common_date}' 행이 {EXCEL_NAME} 2행에 저장되었습니다.")


def main():
    print("[시작] 해외 ETF/지수 데이터 수집 (yfinance)\n")

    common_date, data = fetch_all()

    if not common_date or len(data) < len(TICKERS):
        print("[종료] 완전한 데이터를 수집하지 못했습니다.")
        return

    print(f"\n[수집 성공] 기준 날짜: {common_date}")
    save_to_excel(common_date, data)


if __name__ == "__main__":
    main()