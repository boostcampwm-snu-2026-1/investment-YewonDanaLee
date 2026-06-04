import os
import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

# 매핑할 이름과 인베스팅닷컴 URL (키 이름을 파일명 규칙에 맞추어 수정)
STOCKS = {
    "삼성": "https://kr.investing.com/equities/samsung-electronics-co-ltd-historical-data",
    "SK": "https://kr.investing.com/equities/sk-hynix-inc-historical-data"
}

def create_driver() -> webdriver.Chrome:
    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    )
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    
    driver = webdriver.Chrome(options=opts)
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"},
    )
    return driver

def fetch_closed_market_row(driver, url):
    """실행일 기준 당일 데이터를 제외하고, 완벽히 마감된 '가장 최근 전날 장날' 데이터를 추출"""
    driver.get(url)
    try:
        table = WebDriverWait(driver, 12).until(
            EC.presence_of_element_located((By.XPATH, '//table'))
        )
        
        # 시스템 상의 오늘 날짜 포맷 생성 (예: "2026년 06월 04일")
        today_str = datetime.now().strftime("%Y년 %m월 %d일")
        
        target_tr = None
        for i in [1, 2]:
            tr = table.find_element(By.XPATH, f'.//tbody/tr[{i}]')
            row_date = tr.find_element(By.TAG_NAME, 'time').text.strip()
            
            # 첫 번째 행이 오늘 날짜라면 장 중이므로 제외하고 다음 행(마감된 전날 장날)을 선택
            if row_date == today_str:
                print(f"    [안내] 첫 번째 행이 오늘 날짜({row_date})이므로 마감 전 데이터로 간주해 제외합니다.")
                continue
            
            target_tr = tr
            break
            
        if not target_tr:
            return None
            
        date_final = target_tr.find_element(By.TAG_NAME, 'time').text.strip()
        tds = target_tr.find_elements(By.TAG_NAME, 'td')
        
        return {
            "date": date_final,
            "close": tds[1].text.strip(),
            "open": tds[2].text.strip(),
            "high": tds[3].text.strip(),
            "low": tds[4].text.strip(),
            "volume": tds[5].text.strip(),
            "change": tds[6].text.strip()
        }
    except Exception as e:
        print(f"    [에러] 데이터 수집 중 오류: {e}")
        return None

def save_to_excel(prefix, data):
    # 요청하신 대로 파일명 설정 (_history.xlsx)
    data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data")
    os.makedirs(data_dir, exist_ok=True)
    filename = os.path.join(data_dir, f"{prefix}_history.xlsx")
    
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "시트1"
        ws.append(["날짜", "종가", "시가", "고가", "저가", "거래량", "변동 %"])
        wb.save(filename)
        
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # 중복 저장 방지 검증 (2행의 날짜와 비교)
    if ws.max_row >= 2:
        existing_date = ws.cell(row=2, column=1).value
        if existing_date == data["date"]:
            print(f"  👉 [스킵] {filename}에 {data['date']} 데이터가 이미 존재합니다.\n")
            wb.close()
            return
            
    # 2행에 새로운 빈 행 삽입하여 밀어내기
    ws.insert_rows(2)
    
    # 데이터 매핑
    ws.cell(row=2, column=1, value=data["date"])
    ws.cell(row=2, column=2, value=data["close"])
    ws.cell(row=2, column=3, value=data["open"])
    ws.cell(row=2, column=4, value=data["high"])
    ws.cell(row=2, column=5, value=data["low"])
    ws.cell(row=2, column=6, value=data["volume"])
    ws.cell(row=2, column=7, value=data["change"])
    
    wb.save(filename)
    wb.close()
    print(f"  👉 [성공] {data['date']} 데이터가 {filename}의 2행에 저장되었습니다.\n")

def main():
    driver = create_driver()
    print("[시작] 국내 주식 전일 마감 데이터 수집을 시작합니다.\n")
    
    for prefix, url in STOCKS.items():
        print(f"[{prefix}_history] 데이터 수집 시도 중...")
        stock_data = fetch_closed_market_row(driver, url)
        
        if stock_data:
            print(f"    확정 날짜: {stock_data['date']} (종가: {stock_data['close']} / 거래량: {stock_data['volume']})")
            save_to_excel(prefix, stock_data)
        else:
            print(f"    [실패] {prefix}의 데이터를 파싱하지 못했습니다.\n")
            
    driver.quit()
    print("[완료] 모든 크롤링 프로세스가 종료되었습니다.")

if __name__ == "__main__":
    main()